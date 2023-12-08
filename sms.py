import email, smtplib, ssl
from providers import PROVIDERS
from openpyxl import Workbook, load_workbook

# used for MMS
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from os.path import basename

def send_sms_via_email(
        number: str,
        message: str,
        provider: str,
        sender_credentials: tuple,
        subject: str = "UCSD MASA Reminder Notifications",
        smtp_server: str = "smtp.gmail.com",
        smtp_port: int = 465,
):
    sender_email, email_password = sender_credentials
    receiver_email = f'{number}@{PROVIDERS.get(provider).get("sms")}'

    email_message = f"Subject:{subject}\nTo:{receiver_email}\n{message}"

    with smtplib.SMTP_SSL(
            smtp_server, smtp_port, context=ssl.create_default_context()
    ) as email:
        email.login(sender_email, email_password)
        email.sendmail(sender_email, receiver_email, email_message)

def send_mms_via_email(
        number: str,
        message: str,
        file_path: str,
        mime_maintype: str,
        mime_subtype: str,
        provider: str,
        sender_credentials: tuple,
        subject: str = "sent using etext",
        smtp_server: str = "smtp.gmail.com",
        smtp_port: int = 465,
):

    sender_email, email_password = sender_credentials
    receiver_email = f'{number}@{PROVIDERS.get(provider).get("sms")}'

    email_message=MIMEMultipart()
    email_message["Subject"] = subject
    email_message["From"] = sender_email
    email_message["To"] = receiver_email

    email_message.attach(MIMEText(message, "plain"))

    with open(file_path, "rb") as attachment:
        part = MIMEBase(mime_maintype, mime_subtype)
        part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={basename(file_path)}",
        )

        email_message.attach(part)

    text = email_message.as_string()

    with smtplib.SMTP_SSL(
            smtp_server, smtp_port, context=ssl.create_default_context()
    ) as email:
        email.login(sender_email, email_password)
        email.sendmail(sender_email, receiver_email, text)


def find_phonenumbers(
        excel_file_name: str
):
    wb = load_workbook(excel_file_name)
    ws = wb.active
    x=0
    while x < len(ws['1']):
        x=x+1
        title = ws[chr(64+x)+str(1)].value
        if title =='phone number':
            break
    y=1
    phonenumbers=[];
    while y < len(ws[chr(64+x)]):
        y=y+1
        phonenumbers.append(ws[chr(64+x)+str(y)].value)
    return phonenumbers


def main():
    sender_credentials = ("ucsdmasa@gmail.com", "rqxo yngw lqij ggep")
    message = "Reminder that Event tomorrow"
    excel_file_name = "test.xlsx"
    #input desired message
    # MMS
    # file_path = ""
    # /Users/jazz/Desktop/masa.png
    mime_maintype = "image"
    mime_subtype = "png"
    #input excel file name
    number = find_phonenumbers (excel_file_name)
    # print(pho_numbers)
    counter1=0
    while counter1 < len(number):
        print("Total Percent: "+str((counter1+1)/len(number)*100)+'%')
        counter2 = 0
        while counter2 < 17:
            counter2=counter2+1
            provider = str(counter2)
            print("Sending to User "+str((counter1+1))+": "+str(round((float(provider)/19)*100,2))+'%')
            send_sms_via_email(
                number[counter1],
                message, provider,
                sender_credentials
            )

        counter1=counter1+1

    # SMS

    #MMS
    # send_mms_via_email(
    #     number,
    #     message,
    #     file_path,
    #     mime_maintype,
    #     mime_subtype,
    #     provider,
    #     sender_credentials,
    # )
if __name__ == "__main__":
    main()
