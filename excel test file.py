from openpyxl import Workbook, load_workbook

def find_phonenumbers(
        excel_file_name: str
):
    wb = load_workbook(excel_file_name)
    ws = wb.active
    x=0
    while x < len(ws['1']):
        x=x+1
        title = ws[chr(64+x)+str(1)].value
        if title =='phone number':
            break
    y=1
    phonenumbers=[];
    print(x)
    print(len(ws[chr(64+x)]))
    print(len(ws['B']))
    while y < len(ws[chr(64+x)]):
        y=y+1
        phonenumbers.append(ws[chr(64+x)+str(y)].value)
    return phonenumbers


message = find_phonenumbers('test.xlsx')
print(message[1])
